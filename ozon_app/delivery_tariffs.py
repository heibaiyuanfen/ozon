from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Iterable


_CLUSTER_ALIASES = {
    "воронеж": "沃罗涅日", "voronezh": "沃罗涅日",
    "дальний восток": "远东", "far east": "远东",
    "екатеринбург": "叶卡捷琳堡", "yekaterinburg": "叶卡捷琳堡",
    "казань": "喀山", "kazan": "喀山",
    "калининград": "加里宁格勒", "kaliningrad": "加里宁格勒",
    "краснодар": "克拉斯诺达尔", "krasnodar": "克拉斯诺达尔",
    "красноярск": "克拉斯诺亚尔斯克", "krasnoyarsk": "克拉斯诺亚尔斯克",
    "махачкала": "马哈奇卡拉", "makhachkala": "马哈奇卡拉",
    "москва, мо и дальние регионы": "莫斯科、莫斯科州和周边地区",
    "москва и московская область": "莫斯科、莫斯科州和周边地区",
    "moscow": "莫斯科、莫斯科州和周边地区",
    "нижний новгород": "下诺夫哥罗德", "nizhny novgorod": "下诺夫哥罗德",
    "новосибирск": "新西伯利亚", "novosibirsk": "新西伯利亚",
    "омск": "鄂木斯克", "omsk": "鄂木斯克",
    "оренбург": "奥伦堡", "orenburg": "奥伦堡",
    "пермь": "彼尔姆", "perm": "彼尔姆", "二叠纪": "彼尔姆",
    "ростов": "罗斯托夫", "rostov": "罗斯托夫",
    "самара": "萨马拉", "samara": "萨马拉",
    "санкт-петербург и сзфо": "圣彼得堡和西北地区",
    "санкт-петербург": "圣彼得堡和西北地区", "saint petersburg": "圣彼得堡和西北地区",
    "саратов": "萨拉托夫", "saratov": "萨拉托夫",
    "тверь": "特维尔", "tver": "特维尔",
    "тюмень": "秋明", "tyumen": "秋明",
    "уфа": "乌法", "ufa": "乌法",
    "ярославль": "雅罗斯拉夫尔", "yaroslavl": "雅罗斯拉夫尔",
    "алматы": "阿拉木图", "almaty": "阿拉木图",
    "армения": "亚美尼亚", "armenia": "亚美尼亚",
    "астана": "阿斯塔纳", "astana": "阿斯塔纳",
    "беларусь": "白俄罗斯", "belarus": "白俄罗斯",
    "азербайджан": "阿塞拜疆", "azerbaijan": "阿塞拜疆",
    "грузия": "格鲁吉亚", "georgia": "格鲁吉亚",
    "киргизия": "吉尔吉斯斯坦", "кыргызстан": "吉尔吉斯斯坦",
    "kyrgyzstan": "吉尔吉斯斯坦",
    "узбекистан": "乌兹别克斯坦", "uzbekistan": "乌兹别克斯坦",
    "туркменистан": "土库曼斯坦", "turkmenistan": "土库曼斯坦",
}


def normalize_cluster(value: Any) -> str:
    text = re.sub(r"\s+", " ", str(value or "").strip())
    if not text:
        return ""
    folded = text.casefold().replace("ё", "е")
    if folded in _CLUSTER_ALIASES:
        return _CLUSTER_ALIASES[folded]
    for key, translated in _CLUSTER_ALIASES.items():
        if len(key) >= 5 and key in folded:
            return translated
    return text


def parse_volume_range(value: Any) -> tuple[float, float] | None:
    text = str(value or "").strip().replace(",", ".")
    numbers = re.findall(r"\d+(?:\.\d+)?", text)
    if len(numbers) < 2:
        return None
    low, high = float(numbers[0]), float(numbers[1])
    return (min(low, high), max(low, high))


def read_tariff_workbook(path: str | Path) -> list[dict[str, Any]]:
    """Read the user-provided Ozon route/volume tariff workbook.

    The first sheet contains one row per origin, destination and litre range.
    Column E is the price for goods up to 300 RUB and F is the price above it.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as error:  # pragma: no cover - dependency is packaged
        raise RuntimeError("缺少 openpyxl，无法读取 Ozon 配送费 Excel。") from error
    source = Path(path)
    if not source.is_file():
        raise FileNotFoundError(f"配送费表不存在：{source}")
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        result: list[dict[str, Any]] = []
        for row in sheet.iter_rows(min_row=4, values_only=True):
            volume = parse_volume_range(row[1] if len(row) > 1 else None)
            origin = normalize_cluster(row[2] if len(row) > 2 else "")
            destination = normalize_cluster(row[3] if len(row) > 3 else "")
            if not volume or not origin or not destination:
                continue
            try:
                low_price = float(row[4])
                high_price = float(row[5])
            except (IndexError, TypeError, ValueError):
                continue
            result.append({
                "origin": origin,
                "destination": destination,
                "volume_min_l": volume[0],
                "volume_max_l": volume[1],
                "price_le_300": low_price,
                "price_gt_300": high_price,
            })
        if not result:
            raise ValueError("Excel 中没有识别到配送费路线；请确认使用的是 Ozon 集群配送表。")
        return result
    finally:
        workbook.close()


def tariff_value(
    rows: Iterable[dict[str, Any]], *, origin: str, destination: str,
    volume_l: float, unit_price_rub: float,
) -> float | None:
    wanted_origin = normalize_cluster(origin)
    wanted_destination = normalize_cluster(destination)
    for row in rows:
        if normalize_cluster(row.get("origin")) != wanted_origin:
            continue
        if normalize_cluster(row.get("destination")) != wanted_destination:
            continue
        if float(row.get("volume_min_l") or 0) <= volume_l <= float(
            row.get("volume_max_l") or 0
        ):
            key = "price_le_300" if unit_price_rub <= 300 else "price_gt_300"
            return float(row.get(key) or 0)
    return None
