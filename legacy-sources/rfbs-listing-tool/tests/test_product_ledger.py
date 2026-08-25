import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table


APP_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(APP_DIR))

from product_ledger import LEDGER_COLUMNS, SHEET_TITLE, TABLE_NAME, upsert_product_ledger
from app import RfbsListingApp


def sample_record(**overrides):
    record = {
        "shop_name": "WXZ01",
        "shop_config_id": "shop-1",
        "listing_mode": "本地新品",
        "offer_id": "WXZ-20260806-1",
        "supplier_1688_url": "https://detail.1688.com/offer/123456.html",
        "ozon_source_url": "https://www.ozon.ru/product/5028964005/",
        "ozon_product_id": "5028964005",
        "title": "Светодиодный светильник",
        "model_name": "WXZ-20260806",
        "price": "585",
        "old_price": "731",
        "currency_code": "RUB",
        "pricing_mode": "FBP",
        "purchase_cost": "100",
        "label_fee": "0",
        "target_roi_percent": "60",
        "actual_roi_percent": "61.25",
        "advertising_percent": "15",
        "cargo_loss_percent": "10",
        "raw_sales_commission_percent": "14",
        "sales_commission_percent": "13",
        "shipping": "73.98",
        "sales_commission": "81.90",
        "logistics_commission": "11.70",
        "advertising": "87.75",
        "cargo_loss": "17.60",
        "profit": "104.13",
        "net_weight_g": "2500",
        "gross_weight_g": "2600",
        "depth_mm": "350",
        "width_mm": "350",
        "height_mm": "280",
        "category_display": "住宅和花园 / 炊具 / 蒸锅",
        "category_id": "17028732",
        "type_id": "970895078",
        "warehouse_id": "1020005011815650",
        "stock": "3",
        "import_task_id": "987654321",
        "status": "上架及 RFBS 库存完成",
    }
    record.update(overrides)
    return record


class ProductLedgerTests(unittest.TestCase):
    def test_creates_formatted_workbook_with_typed_values_and_links(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "产品台账.xlsx"
            result = upsert_product_ledger(path, sample_record())

            self.assertEqual(result["action"], "created")
            self.assertEqual(result["row"], 2)
            workbook = load_workbook(path)
            worksheet = workbook[SHEET_TITLE]
            headers = {cell.value: cell.column for cell in worksheet[1]}

            self.assertEqual(worksheet.freeze_panes, "C2")
            self.assertFalse(worksheet.sheet_view.showGridLines)
            self.assertIn(TABLE_NAME, worksheet.tables)
            self.assertEqual(worksheet.cell(2, headers["货号"]).value, "WXZ-20260806-1")
            self.assertEqual(worksheet.cell(2, headers["上品模式"]).value, "本地新品")
            self.assertEqual(worksheet.cell(2, headers["售价"]).value, 585)
            self.assertEqual(worksheet.cell(2, headers["核价模式"]).value, "FBP")
            self.assertEqual(worksheet.cell(2, headers["贴单费"]).value, 0)
            self.assertAlmostEqual(worksheet.cell(2, headers["目标ROI"]).value, 0.6)
            self.assertAlmostEqual(worksheet.cell(2, headers["广告费率"]).value, 0.15)
            self.assertAlmostEqual(worksheet.cell(2, headers["货损率"]).value, 0.10)
            self.assertAlmostEqual(worksheet.cell(2, headers["填写标准佣金"]).value, 0.14)
            self.assertAlmostEqual(worksheet.cell(2, headers["核价佣金"]).value, 0.13)
            self.assertEqual(worksheet.cell(2, headers["目标ROI"]).number_format, "0.00%")
            self.assertEqual(
                worksheet.cell(2, headers["1688采购链接"]).hyperlink.target,
                "https://detail.1688.com/offer/123456.html",
            )
            self.assertEqual(worksheet.cell(2, headers["RFBS仓库ID"]).value, "1020005011815650")
            self.assertTrue(worksheet.column_dimensions[
                worksheet.cell(1, headers["店铺配置ID"]).column_letter
            ].hidden)
            workbook.close()

    def test_updates_same_shop_and_offer_without_adding_a_row(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "产品台账.xlsx"
            upsert_product_ledger(path, sample_record())
            result = upsert_product_ledger(
                path,
                sample_record(price="620", stock="8", supplier_1688_url=""),
            )

            self.assertEqual(result["action"], "updated")
            workbook = load_workbook(path)
            worksheet = workbook[SHEET_TITLE]
            headers = {cell.value: cell.column for cell in worksheet[1]}
            self.assertEqual(worksheet.max_row, 2)
            self.assertEqual(worksheet.cell(2, headers["售价"]).value, 620)
            self.assertEqual(worksheet.cell(2, headers["可售库存"]).value, 8)
            self.assertIsNone(worksheet.cell(2, headers["1688采购链接"]).value)
            self.assertIsNone(worksheet.cell(2, headers["1688采购链接"]).hyperlink)
            workbook.close()

    def test_adds_new_offer_and_keeps_same_offer_separate_between_shops(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "产品台账.xlsx"
            upsert_product_ledger(path, sample_record())
            second = upsert_product_ledger(path, sample_record(offer_id="WXZ-20260806-2"))
            other_shop = upsert_product_ledger(
                path,
                sample_record(shop_name="WXZ02", shop_config_id="shop-2"),
            )

            self.assertEqual(second["action"], "added")
            self.assertEqual(other_shop["action"], "added")
            workbook = load_workbook(path)
            worksheet = workbook[SHEET_TITLE]
            self.assertEqual(worksheet.max_row, 4)
            self.assertEqual(
                worksheet.tables[TABLE_NAME].ref,
                f"A1:{worksheet.cell(1, len(LEDGER_COLUMNS)).column_letter}4",
            )
            workbook.close()

    def test_upgrades_existing_ledger_headers_and_table_for_fbp_columns(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "产品台账.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = SHEET_TITLE
            old_columns = [
                column for column in LEDGER_COLUMNS
                if column.key not in {
                    "listing_mode", "pricing_mode",
                    "advertising_percent", "cargo_loss_percent",
                    "target_profit_margin_percent", "actual_profit_margin_percent",
                    "discount_percent", "cny_rub_rate", "cross_border_shipping_cny",
                    "first_mile_shipping_cny", "warehouse_operation_fee_cny",
                    "tail_delivery_rub", "volume_liters",
                }
            ]
            old_headers = [
                "RFBS销售佣金" if column.key == "sales_commission_percent"
                else "Ozon原始RFBS佣金" if column.key == "raw_sales_commission_percent"
                else column.title
                for column in old_columns
            ]
            worksheet.append(old_headers)
            row = [None] * len(old_headers)
            row[old_headers.index("货号")] = "WXZ-20260806-1"
            row[old_headers.index("店铺配置ID")] = "shop-1"
            worksheet.append(row)
            worksheet.add_table(Table(
                displayName=TABLE_NAME,
                ref=f"A1:{worksheet.cell(1, len(old_headers)).column_letter}2",
            ))
            workbook.save(path)
            workbook.close()

            upsert_product_ledger(path, sample_record())

            upgraded = load_workbook(path)
            worksheet = upgraded[SHEET_TITLE]
            headers = [cell.value for cell in worksheet[1]]
            self.assertIn("核价模式", headers)
            self.assertIn("填写标准佣金", headers)
            self.assertIn("核价佣金", headers)
            self.assertIn("目标利润率", headers)
            self.assertIn("WB尾程费(RUB)", headers)
            self.assertNotIn("RFBS销售佣金", headers)
            self.assertNotIn("Ozon原始RFBS佣金", headers)
            self.assertEqual(len(headers), len(LEDGER_COLUMNS))
            self.assertEqual(
                worksheet.tables[TABLE_NAME].ref,
                f"A1:{worksheet.cell(1, len(LEDGER_COLUMNS)).column_letter}2",
            )
            upgraded.close()

    def test_app_builds_ledger_record_from_completed_job_snapshot(self):
        class Value:
            def __init__(self, value):
                self.value = value

            def get(self):
                return self.value

        app = RfbsListingApp.__new__(RfbsListingApp)
        app.vars = {
            "title": Value("Товар"),
            "supplier_1688_url": Value("must-not-use-current-form"),
        }
        app._ozon_shop_for_api = lambda _shop_id=None: {
            "id": "shop-1", "name": "WXZ01",
        }
        job = {
            "inputs": {
                "ozon_shop_id": "shop-1", "ozon_shop_name": "WXZ01",
                "source_url": "https://www.ozon.ru/product/123/",
                "supplier_1688_url": "https://detail.1688.com/offer/456.html",
                "model_name": "MODEL-1", "net_weight": "800",
                "purchase_cost": "15", "label_fee": "2", "target_roi": "60",
                "sales_commission_percent": "14",
                "advertising_percent": "18", "cargo_loss_percent": "6",
                "fbp_pricing": "1",
                "listing_mode": "local",
                "category_display": "类目 A",
            },
        }
        item = {
            "offer_id": "SKU-1", "name": "Товар", "price": "100", "old_price": "0",
            "currency_code": "RUB", "weight": 900, "depth": 60, "width": 110,
            "height": 240, "description_category_id": 10, "type_id": 20,
        }
        pricing = {
            "product_id": "123", "actual_roi_percent": "61.00",
            "pricing_mode": "FBP", "manual_sales_commission_percent": "14",
            "sales_percent_rfbs": "13", "profit": "10.37",
        }

        record = app._product_ledger_record(item, pricing, 99, 1001, 3, job_record=job)

        self.assertEqual(record["offer_id"], "SKU-1")
        self.assertEqual(record["listing_mode"], "本地新品")
        self.assertEqual(record["supplier_1688_url"], "https://detail.1688.com/offer/456.html")
        self.assertEqual(record["ozon_source_url"], "https://www.ozon.ru/product/123/")
        self.assertEqual(record["net_weight_g"], "800")
        self.assertEqual(record["gross_weight_g"], 900)
        self.assertEqual(record["import_task_id"], 99)
        self.assertEqual(record["pricing_mode"], "FBP")
        self.assertEqual(record["label_fee"], "0")
        self.assertEqual(record["advertising_percent"], "18")
        self.assertEqual(record["cargo_loss_percent"], "6")
        self.assertEqual(record["raw_sales_commission_percent"], "14")
        self.assertEqual(record["sales_commission_percent"], "13")

        del job["inputs"]["supplier_1688_url"]
        legacy_record = app._product_ledger_record(item, pricing, 99, 1001, 3, job_record=job)
        self.assertEqual(legacy_record["supplier_1688_url"], "")


if __name__ == "__main__":
    unittest.main()
