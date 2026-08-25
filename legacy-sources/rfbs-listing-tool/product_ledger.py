from __future__ import annotations

import os
import threading
import uuid
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


SHEET_TITLE = "产品台账"
TABLE_NAME = "ProductLedgerTable"


@dataclass(frozen=True)
class LedgerColumn:
    key: str
    title: str
    width: float
    kind: str = "text"


LEDGER_COLUMNS = (
    LedgerColumn("updated_at", "更新时间", 20, "datetime"),
    LedgerColumn("platform", "平台", 14),
    LedgerColumn("shop_name", "上品店铺", 18),
    LedgerColumn("listing_mode", "上品模式", 12),
    LedgerColumn("offer_id", "货号", 22),
    LedgerColumn("supplier_1688_url", "1688采购链接", 38, "link"),
    LedgerColumn("ozon_source_url", "Ozon来源链接", 38, "link"),
    LedgerColumn("ozon_product_id", "Ozon商品ID", 18),
    LedgerColumn("wb_source_url", "WB来源链接", 38, "link"),
    LedgerColumn("wb_source_nm_id", "WB来源商品ID", 18),
    LedgerColumn("wb_nm_id", "WB商品ID(nmID)", 18),
    LedgerColumn("wb_chrt_id", "WB规格ID(chrtID)", 18),
    LedgerColumn("wb_subject_id", "WB子类目ID", 16),
    LedgerColumn("title", "商品标题", 36),
    LedgerColumn("model_name", "型号名称", 24),
    LedgerColumn("price", "售价", 13, "money"),
    LedgerColumn("old_price", "折扣前价格", 13, "money"),
    LedgerColumn("currency_code", "币种", 9),
    LedgerColumn("pricing_mode", "核价模式", 11),
    LedgerColumn("purchase_cost", "采购成本", 13, "money"),
    LedgerColumn("label_fee", "贴单费", 11, "money"),
    LedgerColumn("target_roi_percent", "目标ROI", 11, "percent"),
    LedgerColumn("actual_roi_percent", "实际ROI", 11, "percent"),
    LedgerColumn("target_profit_margin_percent", "目标利润率", 12, "percent"),
    LedgerColumn("actual_profit_margin_percent", "实际利润率", 12, "percent"),
    LedgerColumn("discount_percent", "WB折扣", 10, "percent"),
    LedgerColumn("advertising_percent", "广告费率", 11, "percent"),
    LedgerColumn("cargo_loss_percent", "货损率", 11, "percent"),
    LedgerColumn("raw_sales_commission_percent", "填写标准佣金", 15, "percent"),
    LedgerColumn("sales_commission_percent", "核价佣金", 13, "percent"),
    LedgerColumn("shipping", "运费", 12, "money"),
    LedgerColumn("cny_rub_rate", "CNY兑RUB汇率", 14, "number"),
    LedgerColumn("cross_border_shipping_cny", "WB跨境运费(CNY)", 18, "money"),
    LedgerColumn("first_mile_shipping_cny", "WB头程运费(CNY)", 18, "money"),
    LedgerColumn("warehouse_operation_fee_cny", "海外仓操作费(CNY)", 19, "money"),
    LedgerColumn("tail_delivery_rub", "WB尾程费(RUB)", 16, "money"),
    LedgerColumn("volume_liters", "包装体积(L)", 13, "number"),
    LedgerColumn("sales_commission", "销售佣金", 12, "money"),
    LedgerColumn("logistics_commission", "物流佣金", 12, "money"),
    LedgerColumn("advertising", "广告费", 12, "money"),
    LedgerColumn("cargo_loss", "货损", 12, "money"),
    LedgerColumn("profit", "利润", 12, "money"),
    LedgerColumn("net_weight_g", "商品净重(g)", 14, "number"),
    LedgerColumn("gross_weight_g", "包装毛重(g)", 14, "number"),
    LedgerColumn("depth_mm", "包装长度(mm)", 15, "number"),
    LedgerColumn("width_mm", "包装宽度(mm)", 15, "number"),
    LedgerColumn("height_mm", "包装高度(mm)", 15, "number"),
    LedgerColumn("category_display", "Ozon类目", 34),
    LedgerColumn("category_id", "类目ID", 16),
    LedgerColumn("type_id", "类型ID", 16),
    LedgerColumn("warehouse_id", "RFBS仓库ID", 20),
    LedgerColumn("stock", "可售库存", 11, "integer"),
    LedgerColumn("import_task_id", "Ozon导入任务ID", 18),
    LedgerColumn("status", "状态", 22),
    LedgerColumn("shop_config_id", "店铺配置ID", 18),
)


_LEDGER_LOCK = threading.RLock()
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
_BODY_FONT = Font(name="Microsoft YaHei", size=10, color="1F2933")
_LINK_FONT = Font(name="Microsoft YaHei", size=10, color="0563C1", underline="single")
_THIN_GRAY = Side(style="thin", color="D9E2F3")
_BODY_BORDER = Border(bottom=_THIN_GRAY)


def _safe_text(value) -> str:
    text = str(value or "").strip()
    if text.startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def _decimal_value(value, title: str):
    if value is None or str(value).strip() == "":
        return None
    try:
        number = Decimal(str(value).strip())
    except (InvalidOperation, ValueError) as error:
        raise ValueError(f"产品台账字段“{title}”必须是有效数字：{value}") from error
    if not number.is_finite():
        raise ValueError(f"产品台账字段“{title}”必须是有限数字：{value}")
    return float(number)


def _cell_value(column: LedgerColumn, value):
    if column.kind in {"text", "link"}:
        return _safe_text(value)
    if column.kind == "datetime":
        if isinstance(value, datetime):
            return value.replace(microsecond=0)
        if value is None or str(value).strip() == "":
            return datetime.now().replace(microsecond=0)
        text = str(value).strip()
        try:
            return datetime.fromisoformat(text)
        except ValueError:
            return _safe_text(text)
    if column.kind == "integer":
        number = _decimal_value(value, column.title)
        return int(number) if number is not None else None
    number = _decimal_value(value, column.title)
    if column.kind == "percent" and number is not None:
        return number / 100
    return number


def _load_or_create_workbook(path: Path):
    if not path.exists():
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = SHEET_TITLE
        workbook.properties.title = "Ozon RFBS 产品台账"
        workbook.properties.subject = "自动上架完成记录"
        return workbook, worksheet, True
    try:
        workbook = load_workbook(path)
    except PermissionError as error:
        raise PermissionError(f"无法读取产品台账，请先关闭 Excel：{path}") from error
    except Exception as error:
        raise ValueError(f"无法读取现有产品台账，请检查文件是否损坏：{path}\n{error}") from error
    worksheet = workbook[SHEET_TITLE] if SHEET_TITLE in workbook.sheetnames else workbook.create_sheet(SHEET_TITLE)
    return workbook, worksheet, False


def _ensure_headers(worksheet) -> dict[str, int]:
    header_by_title = {
        str(cell.value).strip(): cell.column
        for cell in worksheet[1]
        if cell.value is not None and str(cell.value).strip()
    }
    legacy_commission_title = "RFBS销售佣金"
    if "核价佣金" not in header_by_title and legacy_commission_title in header_by_title:
        column_index = header_by_title.pop(legacy_commission_title)
        worksheet.cell(1, column_index, "核价佣金")
        header_by_title["核价佣金"] = column_index
    legacy_manual_commission_title = "Ozon原始RFBS佣金"
    if "填写标准佣金" not in header_by_title and legacy_manual_commission_title in header_by_title:
        column_index = header_by_title.pop(legacy_manual_commission_title)
        worksheet.cell(1, column_index, "填写标准佣金")
        header_by_title["填写标准佣金"] = column_index
    for column in LEDGER_COLUMNS:
        if column.title not in header_by_title:
            index = len(header_by_title) + 1
            while worksheet.cell(1, index).value not in (None, ""):
                index += 1
            worksheet.cell(1, index, column.title)
            header_by_title[column.title] = index
    return header_by_title


def _matching_row(worksheet, header_by_title: dict[str, int], record: dict) -> int | None:
    offer_id = str(record.get("offer_id") or "").strip()
    shop_config_id = str(record.get("shop_config_id") or "").strip()
    offer_column = header_by_title["货号"]
    shop_column = header_by_title["店铺配置ID"]
    for row in range(2, worksheet.max_row + 1):
        existing_offer = str(worksheet.cell(row, offer_column).value or "").strip().lstrip("'")
        if existing_offer != offer_id:
            continue
        existing_shop = str(worksheet.cell(row, shop_column).value or "").strip().lstrip("'")
        if shop_config_id and existing_shop and existing_shop != shop_config_id:
            continue
        return row
    return None


def _style_worksheet(worksheet, header_by_title: dict[str, int], target_row: int):
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "C2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(max(header_by_title.values()))}{worksheet.max_row}"
    worksheet.row_dimensions[1].height = 28
    worksheet.row_dimensions[target_row].height = 22
    worksheet.sheet_properties.tabColor = "1F4E78"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True

    for column in LEDGER_COLUMNS:
        column_index = header_by_title[column.title]
        letter = get_column_letter(column_index)
        worksheet.column_dimensions[letter].width = column.width
        header = worksheet.cell(1, column_index)
        header.fill = _HEADER_FILL
        header.font = _HEADER_FONT
        header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row, column_index)
            cell.font = _BODY_FONT
            cell.border = _BODY_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=column.kind == "text")
            if column.kind in {"money", "number"}:
                cell.number_format = "#,##0.00"
            elif column.kind == "percent":
                cell.number_format = "0.00%"
            elif column.kind == "integer":
                cell.number_format = "0"
            elif column.kind == "datetime":
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
            else:
                cell.number_format = "@"
            if column.kind == "link" and str(cell.value or "").startswith(("http://", "https://")):
                cell.hyperlink = str(cell.value)
                cell.font = _LINK_FONT
            elif column.kind == "link":
                cell.hyperlink = None

    worksheet.column_dimensions[get_column_letter(header_by_title["店铺配置ID"])].hidden = True


def _ensure_table(worksheet, header_by_title: dict[str, int]):
    if worksheet.max_row < 2:
        return
    last_column = max(header_by_title.values())
    for index in range(1, last_column + 1):
        if worksheet.cell(1, index).value in (None, ""):
            worksheet.cell(1, index, f"字段{index}")
    reference = f"A1:{get_column_letter(last_column)}{worksheet.max_row}"
    if worksheet.tables.get(TABLE_NAME) is not None:
        del worksheet.tables[TABLE_NAME]
    table = Table(displayName=TABLE_NAME, ref=reference)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def _save_atomic(workbook, path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = path.with_name(f".{path.stem}-{uuid.uuid4().hex}.tmp.xlsx")
    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, path)
    except PermissionError as error:
        raise PermissionError(
            f"无法更新产品台账，请关闭正在打开的 Excel 文件后重试：{path}"
        ) from error
    finally:
        try:
            temporary_path.unlink(missing_ok=True)
        except OSError:
            pass


def upsert_product_ledger(path: str | Path, record: dict) -> dict:
    """Create or update one listing row, keyed by shop and Offer ID."""
    ledger_path = Path(path).expanduser().resolve()
    offer_id = str(record.get("offer_id") or "").strip()
    if not offer_id:
        raise ValueError("写入产品台账时货号不能为空")

    with _LEDGER_LOCK:
        workbook, worksheet, workbook_created = _load_or_create_workbook(ledger_path)
        header_by_title = _ensure_headers(worksheet)
        target_row = _matching_row(worksheet, header_by_title, record)
        action = "updated"
        if target_row is None:
            target_row = max(2, worksheet.max_row + 1)
            action = "created" if workbook_created and target_row == 2 else "added"

        values = dict(record)
        values["updated_at"] = values.get("updated_at") or datetime.now()
        for column in LEDGER_COLUMNS:
            cell = worksheet.cell(target_row, header_by_title[column.title])
            cell.value = _cell_value(column, values.get(column.key))

        _style_worksheet(worksheet, header_by_title, target_row)
        _ensure_table(worksheet, header_by_title)
        _save_atomic(workbook, ledger_path)

    return {
        "path": str(ledger_path),
        "sheet": SHEET_TITLE,
        "row": target_row,
        "action": action,
        "offer_id": offer_id,
    }
